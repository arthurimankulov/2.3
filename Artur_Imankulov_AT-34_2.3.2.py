import csv
import os
from operator import itemgetter
from typing import List, Dict
import re
import doctest
import unittest
import numpy as np
import pandas as pd
import openpyxl
from matplotlib import pyplot as plt
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from jinja2 import Environment, FileSystemLoader
import pdfkit
from xlsx2html import xlsx2html


class Salary:
    """
    Класс для представления зарплат
    Attributes:
        salary_from (str): Нижняя граница оклада
        :type (str or int or float)
        salary_to: Верхняя граница оклада
        :type (str or int or float)
        salary_currency: Валюта оклада
        :type (str)
    """
    def __init__(self, salary_from : str or int or float, salary_to : str or int or float, salary_currency : str):
        """
        @param salary_from: Нижняя граница оклада
        :type (str or int or float)
        @param salary_to: Верхняя граница оклада
        :type (str or int or float)
        @param salary_currency: Валюта оклада
        :type (str)
        """
        self.salary_from = float(salary_from)
        self.salary_to = float(salary_to)
        self.salary_currency = salary_currency
        self.average_salary = (int(float(salary_from) + float(salary_to)) / 2)

    def convert_to_rubles(self):
        """
        Конверирует из другой валюты в рубли
        :return: Возвращает значение после конвертированя валюты
        >>> Salary(100, 200, "RUR").convert_to_rubles()
        150.0
        >>> Salary(100, 200, "AZN").convert_to_rubles()
        5352.0
        """
        currency_to_rub = {
            "AZN": 35.68,
            "BYR": 23.91,
            "EUR": 59.90,
            "GEL": 21.74,
            "KGS": 0.76,
            "KZT": 0.13,
            "RUR": 1,
            "UAH": 1.64,
            "USD": 60.66,
            "UZS": 0.0055
        }
        return self.average_salary * currency_to_rub[self.salary_currency]


class Vacancy:
    """
    Класс для представления вакансий
    Attributes:
        name: Название
        :type (str)
        salary: Зарплата для данной вакансии
        :type (Salary)
        area_name: Местоположение
        :type (str)
        published_at: Дата публикации
        :type (str)
        year: Год публикации
        :type (str)
    """
    def __init__(self, vacancy: Dict[str, str]):
        """
        @param vacancy: Отдельная вакансия в виде словаря: атрибут - значение
        :type (Dict[str, str])
        >>> vac = {"name" :"Инженер", "salary_from" : 35000.0,"salary_to" : 45000.0, "salary_currency" : "RUR", "area_name" : "Moscow","published_at" :"2007-12-03T17:47:55+0300"}
        >>> vac = Vacancy(vac)
        >>> vac.area_name
        'Moscow'
        >>> vac.year
        '2007'
        """
        self.name = vacancy["name"]
        self.salary = Salary(salary_from=vacancy["salary_from"],
                             salary_to=vacancy["salary_to"],
                             salary_currency=vacancy["salary_currency"])
        self.area_name = vacancy["area_name"]
        self.published_at = vacancy["published_at"]
        self.year = self.published_at[:4]


class DataSet:
    """
    Класс для представления набора вакансий
    Attributes:
        file_name: Название файла
        :type (str)
        vacancies_objects: писок вакансий
        :type (List[Vacancy])
    """
    def __init__(self, file_name : str):
        """
        @param file_name: Название файла
        :type (str)
        """
        self.file_name = file_name
        self.vacancies_objects = self.__csv_reader()

    def __csv_reader(self) -> (List[Vacancy]):
        """
        Читает из csv файла вакансии и возвращает в виде списка вакансий
        @return: Список вакансий
        :type (List[Vacancy])
        """
        with open(self.file_name, encoding='utf-8-sig') as file:
            file_reader = csv.reader(file)
            lines = [row for row in file_reader]
            headlines, vacancies = lines[0], lines[1:]
        return self.process_vacancies(headlines, vacancies)

    def process_vacancies(self, headlines : List[str], vacancies : List[List[str]]) -> (List[Vacancy]):
        """
        Отбирает правильно заполненные вакансии и конвертирует в класс Vacancy
        :param headlines: Названия заголовков
        :type (List[str])
        :param vacancies: Список из списокв вакансий
        :type (List[List[str]])
        :return: Правильно заполненные вакансии
        :type (List[Vacancy])
        """
        result = []
        for vacancy in vacancies:
            if (len(vacancy) == len(headlines)) and (all([v != "" for v in vacancy])):
                vacancy = [" ".join(re.sub("<.*?>", "", value).replace('\n', '; ').split()) for value in vacancy]
                vacancy = {x: y for x, y in zip([r for r in headlines], [v for v in vacancy])}
                vacancy = Vacancy(vacancy)
                result.append(vacancy)
        return result


class ParamSalary:
    """
    Класс для представления параметра и связанной с ним зарплаты
    Attributes:
        param: Выбранный параметр соотносимый с определенной зарплатой
        :type (str)
        salary: Зарплата для определенной вакансии
        :type (Salary)
        currency_to_rub: Таблица перевода из определенной валюты в рубли
        :type (Dict[str, float])
        count_vacancies: Количество вакансий:
        :type (int)
    """
    def __init__(self, param : str, salary: Salary):
        """
        @param param: Выбранный параметр соотносимый с определенной зарплатой
        :type (str)
        @param salary: Зарплата для определенной вакансии
        :type (Salary)
        >>> ParamSalary("year", Salary(1000, 2000, "USD")).salary
        90990.0
        """
        self.param = param
        self.salary = salary.convert_to_rubles()
        self.count_vacancies = 1

    def add_salary(self, new_salary : Salary):
        """
        Добавление зарплаты к выбранному параметру
        @param new_salary: Зарплата для добавления
        :type (Salary)
        @return: None
        >>> ParamSalary("city", Salary(100, 200, "RUR")).add_salary(Salary(200, 300, "RUR")).salary
        400.0
        >>> ParamSalary("city", Salary(100, 200, "RUR")).add_salary(Salary(200, 300, "RUR")).count_vacancies
        2
        """
        self.count_vacancies += 1
        self.salary = self.salary + new_salary.convert_to_rubles()
        return self


class Report:
    """
    Класс для создания Excel таблицы
    Attributes:
        profession: Профессия
        :type (str)
        years: Года представленных данных
        :type (str)
        average_salary: Набор средних зарплат по годам
        :type (List[int])
        average_salary_profession: Набор средних зарплат по годам для выбранной профессии
        :type (List[int])
        count_vacancies_by_year: Набор количеств вакансий за год
        :type (List[int])
        count_vacancies_by_year_prof: Набор количеств вакансий за год для выбранной професии
        :type (List[int])
        city_salary: Словарь с предмтавлением: город - зарплата
        :type (Dict[str, int])
        city_count_vacancies: Словарь с предмтавлением: город - количество вакансий
        :type (Dict[str, int])
        file_name: Название файла
        :type (str)
    """
    def __init__(self, profession : str, years: List[int], average_salary : List[int],
                 average_salary_profession : List[int], count_vacancies_by_year : List[int],
                 count_vacancies_by_year_prof : List[int], city_salary : Dict[str, int],
                 city_count_vacancies : Dict[str, int], file_name : str):
        """
        @param profession: Профессия
        :type (str)
        @param years: Года представленных данных
        :type (List[int])
        @param average_salary: Набор средних зарплат по годам
        :type (List[int])
        @param average_salary_profession: Набор средних зарплат по годам для выбранной профессии
        :type (List[int])
        @param count_vacancies_by_year: Набор количеств вакансий за год
        :type (List[int])
        @param count_vacancies_by_year_prof: Набор количеств вакансий за год для выбранной професии
        :type (List[int])
        @param city_salary: Словарь с предмтавлением: город - зарплата
        :type (Dict[str, int])
        @param city_count_vacancies: Словарь с предмтавлением: город - количество вакансий
        :type (Dict[str, int])
        @param file_name: Название файла
        :type (str)
        """
        self.years = years
        self.average_salary = average_salary
        self.average_salary_profession = average_salary_profession
        self.count_vacancies_by_year = count_vacancies_by_year
        self.count_vacancies_by_year_prof = count_vacancies_by_year_prof
        self.city_salary = city_salary
        self.city_vacancies = city_count_vacancies
        self.profession = profession
        self.file_name = file_name

    def generate_excel(self) -> None:
        """
        Генерирует файл для полученных данных
        @return: None
        """
        if not isinstance(self.file_name, str):
            raise TypeError('')
        if os.path.basename(self.file_name).split('.')[1] != "xlsx":
            raise TypeError('')
        if os.path.exists(self.file_name):
            raise FileExistsError("")
        df = [[self.years[i], self.average_salary[i], self.average_salary_profession[i], self.count_vacancies_by_year[i], self.count_vacancies_by_year_prof[i]] for i in range(len(self.years))]
        df.insert(0, ["Год", "Средняя зарплата", f"Средняя зарплата - {self.profession}", "Количество вакансий", f"Количество вакансий - {self.profession}"])
        df = pd.DataFrame(df, columns=None)
        cities_of_salary, salaries = [city for city in self.city_salary], [self.city_salary[city] for city in self.city_salary]
        cities_of_vacancy, vacancies = [city for city in self.city_vacancies], ['{:.2f}'.format(self.city_vacancies[city] * 100) + "%" for city in self.city_vacancies]
        df2 = [[cities_of_salary[i], salaries[i], "", cities_of_vacancy[i], vacancies[i]] for i in range(len(cities_of_salary))]
        df2.insert(0, ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"])
        df2 = pd.DataFrame(df2, columns=None)
        with pd.ExcelWriter(self.file_name) as writer:
            df.to_excel(writer, sheet_name='Статистика по годам', index=False, header=False)
            df2.to_excel(writer, sheet_name="Статистика по городам", index=False, header=False)
        wb = openpyxl.load_workbook(self.file_name)
        worksheet1 = wb["Статистика по годам"]
        worksheet2 = wb["Статистика по городам"]
        thin = Side(border_style="thin")
        self.__add_border_and_align(worksheet1, thin, len(self.years) + 2, ["A", "B", "C", "D", "E"])
        self.__add_border_and_align(worksheet2, thin, max(len(cities_of_salary) + 2, len(cities_of_vacancy) + 2), ["A", "B", "D", "E"])
        self.__make_max_column_width(worksheet1)
        self.__make_max_column_width(worksheet2)
        wb.save(self.file_name)

    def __add_border_and_align(self, worksheet : Worksheet, side : Side, count_columns : int, rows : List[str]) -> None:
        """
        @param worksheet: Рабочий лист
        :type (Worksheet)
        @param side: Сторона печати
        :type (Side)
        @param count_columns: Количество столбцов
        :type (int)
        @param rows: Список выбранных названий строк
        :type (List[str])
        @return: None
        """
        for i in range(1, count_columns):
            for row in rows:
                if i == 1:
                    worksheet[row + str(i)].alignment = Alignment(horizontal='left')
                    worksheet[row + str(i)].font = Font(bold=True)
                if worksheet[row + str(i)].internal_value != None:
                    worksheet[row + str(i)].border = Border(top=side, bottom=side, left=side, right=side)

    def __make_max_column_width(self, worksheet : Worksheet) -> None:
        """
        @param worksheet: Рабочий лист
        :type (Worksheet)
        @return: None
        """
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value != None:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                else:
                    dims[cell.column] = len(str(cell.value))
        for col, value in dims.items():
            worksheet.column_dimensions[get_column_letter(col)].width = value + 2


class Graphic:
    """
    Класс для создания графиков с помощью библиотеки matpolib
    Attributes:
        profession: Профессия
        :type (str)
        years: Года представленных данных
        :type (str)
        average_salary: Набор средних зарплат по годам
        :type (List[int])
        average_salary_profession: Набор средних зарплат по годам для выбранной профессии
        :type (List[int])
        count_vacancies_by_year: Набор количеств вакансий за год
        :type (List[int])
        count_vacancies_by_year_prof: Набор количеств вакансий за год для выбранной професии
        :type (List[int])
        city_salary: Словарь с предмтавлением: город - зарплата
        :type (Dict[str, int])
        city_count_vacancies: Словарь с предмтавлением: город - количество вакансий
        :type (Dict[str, int])
        file_name: Название файла
        :type (str)
    """
    def __init__(self, profession: str, years: List[int], average_salary: List[int],
                 average_salary_profession: List[int], count_vacancies_by_year: List[int],
                 count_vacancies_by_year_prof: List[int], city_salary: Dict[str, int], city_vacancies: Dict[str, int],
                 file_name : str):
        """
        @param profession: Выбранная профессия
        :type (str)
        @param years: Года представленных данных
        :type (List[int])
        @param average_salary: Набор средних зарплат по годам
        :type (List[int])
        @param average_salary_profession: Набор средних зарплат по годам для выбранной профессии
        :type (List[int])
        @param count_vacancies_by_year: Набор количеств вакансий за год
        :type (List[int])
        @param count_vacancies_by_year_prof: Набор количеств вакансий за год для выбранной професии
        :type (List[int])
        @param city_salary: Словарь с предмтавлением: город - зарплата
        :type (Dict[str, int])
        @param city_count_vacancies: Словарь с предмтавлением: город - количество вакансий
        :type (Dict[str, int])
        @param file_name: Название файла
        :type (str)
        """
        if not isinstance(file_name, str):
            raise TypeError('')
        if os.path.basename(file_name).split('.')[1] != "png":
            raise TypeError('')
        if os.path.exists(file_name):
            raise FileExistsError("")
        self.years = years
        self.average_salary = average_salary
        self.average_salary_profession = average_salary_profession
        self.count_vacancies_by_year = count_vacancies_by_year
        self.count_vacancies_by_year_prof = count_vacancies_by_year_prof
        self.city_salary = city_salary
        self.city_vacancies = city_vacancies
        self.profession = profession
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2, figsize=(12, 8))
        self.__grouped_bar_graph(ax1, "Уровень зарплат по годам", self.average_salary, self.years,
                                 self.average_salary_profession, 'средняя з/п', f'з/п {self.profession}')
        self.__grouped_bar_graph(ax2, 'Количество вакансий по годам', self.count_vacancies_by_year, self.years,
                                 self.count_vacancies_by_year_prof, 'Количество вакансий',
                                 f'Количество вакансий {self.profession}')
        self.__horizontal_graph(ax3)
        self.__pie_graph(ax4)
        plt.tight_layout()
        #plt.show()
        fig.savefig(file_name)

    def __grouped_bar_graph(self, ax, title: str, values_x: List[int], values_y: List[int], values_x2: List[int],
                            label_x: str, label_x2: str) -> None:
        """
        Создание сгруппированной гистограммы
        @param ax: Местоположение графика
        :type (matplotlib.axes._subplots.AxesSubplot)
        @param title: Название графика
        :type (str)
        @param values_x: Значения 1 выборки по оси X
        :type (List[int])
        @param values_y: Значения  выборки по оси Y
        :type (List[int])
        @param values_x2: Значения 2 выборки по оси X
        :type (List[int])
        @param label_x: Название легенды для 1 выборки
        :type (str)
        @param label_x2: Название легенды для 2 выборки
        :type (str)
        @return: None
        """
        ax.grid(axis='y')
        x = np.arange(len(values_y))
        width = 0.4
        ax.bar(x - width / 2, values_x, width, label=label_x)
        ax.bar(x + width / 2, values_x2, width, label=label_x2)
        ax.set_xticks(x, values_y, rotation=90)
        ax.tick_params(axis="both", labelsize=16)
        ax.set_title(title, fontweight='normal', fontsize=20)
        ax.legend(loc="upper left", fontsize=14)

    def __horizontal_graph(self, ax) -> None:
        """
        Создание горизонтального графика
        @param ax: Местоположение графика
        :type (matplotlib.axes._subplots.AxesSubplot)
        @return: None
        """
        ax.grid(axis='x')
        plt.rcdefaults()
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(10)
        city_salary = ["\n".join(city.split(" ")) for city in self.city_salary]
        ax.barh([city for city in city_salary], [self.city_salary[key] for key in self.city_salary], align='center')
        ax.invert_yaxis()
        ax.set_title('Уровень зарплат по городам', fontweight='normal',  fontsize=20)

    def __pie_graph(self, ax) -> None:
        """
        Создание круговой диаграммы
        @param ax: Местоположение графика
        :type (matplotlib.axes._subplots.AxesSubplot)
        @return: None
        """
        plt.style.use('_mpl-gallery-nogrid')
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(16)
        vacancies = [self.city_vacancies[v] * 100 for v in self.city_vacancies]
        cities = [city for city in self.city_vacancies]
        sum_vacancies = sum(vacancies)
        if sum_vacancies != 100:
            vacancies.insert(0, 100 - sum_vacancies)
            cities.insert(0, "Другие")
        ax.set_title('Доля вакансий по городам', fontweight='normal',  fontsize=20)
        ax.pie(vacancies, labels=cities)


class PdfConverter:
    """
    Класс для конвертирования данных статистики в pdf-файл
    Attributes:
        graph_name: Название файла с графиком (путь относительный в той же папке, что и код)
        :type (str)
        excel_file_name: Название файла с excel таблицой (путь относительный в той же папке, что и код)
        :type (str)
        profession: Название профессии
        :type (str)
        request: Запрос (Вакансии или Статистика)
        :type (str)
    """
    def __init__(self, graph_name : str, excel_file_name : str, profession : str, request : str):
        """
        @param graph_name: Название файла с графиком (путь относительный в той же папке, что и код)
        :type (str)
        @param excel_file_name: Название файла с excel таблицой (путь относительный в той же папке, что и код)
        :type (str)
        @param profession: Название профессии
        :type (str)
        @param request: Запрос (Вакансии или Статистика)
        :type (str)
        """
        self.graph = graph_name
        self.excel_file = excel_file_name
        self.prof = profession
        self.request = request

    def generate_pdf(self) -> None:
        """
        Генерирует pdf-файл из представленных данных
        @return: None
        """
        env = Environment(loader=FileSystemLoader('.'))
        if self.request == "Вакансии" :
            template = env.get_template("pdf_graph.html")
            graph_path = os.path.abspath(self.graph)
            pdf_template = template.render({"prof": self.prof,
                                            "graph": graph_path})
        elif self.request == "Статистика":
            template = env.get_template("pdf_excel.html")
            out_stream2 = xlsx2html(self.excel_file, sheet="Статистика по городам")
            out_stream2.seek(0)
            out_stream = xlsx2html(self.excel_file, sheet="Статистика по годам")
            out_stream.seek(0)
            pdf_template = template.render({"prof" : self.prof,
                                            "first_table" : out_stream.read(),
                                            "second_table" : out_stream2.read()})
        else:
            raise TypeError("")
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


class InputConnect:
    """
    Класс для обработки и иницилизации данных
    Attributes:
        input_data: Данные представленные пользователем (Запрос, имя файла, название нужной профессии)
        :type (List[str])
    """
    def __init__(self):
        """
        Иницилизация данных
        """
        self.input_data = []
        for question in ["Введите запрос: ", "Введите название файла: ", "Введите название профессии: "]:
            print(question, end="")
            self.input_data.append(input())

    def process_data(self) -> None:
        """
        Обработка данных
        @return: None
        >>> dataSet = DataSet("32.txt").process_vacancies(["name","salary_from",'salary_to','salary_currency','area_name',"published_at"],[["Системный аналитик","75000.0",'95000.0','RUR','Москва','2007-12-03T17:41:49+0300']])[0]
        >>> dataSet.name
        'Системный аналитик'
        >>> dataSet.year
        '2007'
        """
        request = self.input_data[0]
        data = DataSet(self.input_data[1]).vacancies_objects
        data_profession = [d for d in data if self.input_data[2] in d.name]
        year_salary = self.convert_to_param_salary(data, "year")
        cities_salary = self.convert_to_param_salary(data, "city")
        professions_year_salary = self.__add_missing_years(self.convert_to_param_salary(data_profession, "year"), year_salary)
        city_salary = dict(sorted({x: y for x, y in zip([r.param for r in cities_salary], [int(v.salary / v.count_vacancies) for v in cities_salary])}.items(), key=itemgetter(1), reverse=True))
        city_vacancies = dict(sorted({x: y for x, y in zip([r.param for r in cities_salary], [v.count_vacancies / len(data) for v in cities_salary])}.items(), key=itemgetter(1), reverse=True))
        year_salary, year_vacancy = self.__convert_from_param_salary_to_dict(year_salary)
        professions_year_salary, professions_year_vacancies = self.__convert_from_param_salary_to_dict(professions_year_salary)
        city_salary = {x : y for x, y in zip([key for key in city_salary if city_vacancies[key] >= 0.01][:10], [city_salary[key] for key in city_salary if city_vacancies[key] >= 0.01])}
        city_vacancies = {x : y for x, y in zip([key for key in city_vacancies if city_vacancies[key] >= 0.01][:10], [float('{:.4f}'.format(city_vacancies[key])) for key in city_vacancies if city_vacancies[key] >= 0.01])}
        output_data = { "Динамика уровня зарплат по годам:" : year_salary,
                        "Динамика количества вакансий по годам:" : year_vacancy,
                        "Динамика уровня зарплат по годам для выбранной профессии:" : professions_year_salary,
                        "Динамика количества вакансий по годам для выбранной профессии:" : professions_year_vacancies,
                        "Уровень зарплат по городам (в порядке убывания):" : city_salary,
                        "Доля вакансий по городам (в порядке убывания):" : city_vacancies}
        [print(i, output_data[i]) for i in output_data]
        excel_file = "report.xlsx"
        profession = self.input_data[2]
        report = Report(profession=profession,
                        years=[i for i in year_salary],
                        average_salary=[year_salary[i] for i in year_salary],
                        average_salary_profession=[professions_year_salary[i] for i in professions_year_salary],
                        count_vacancies_by_year=[year_vacancy[i] for i in year_vacancy],
                        count_vacancies_by_year_prof=[professions_year_vacancies[i] for i in professions_year_vacancies],
                        city_salary=city_salary,
                        city_count_vacancies=city_vacancies,
                        file_name=excel_file)
        report.generate_excel()
        graph_name = "graph.png"
        graph = Graphic(profession=profession,
                        years=[i for i in year_salary],
                        average_salary=[year_salary[i] for i in year_salary],
                        average_salary_profession=[professions_year_salary[i] for i in professions_year_salary],
                        count_vacancies_by_year=[year_vacancy[i] for i in year_vacancy],
                        count_vacancies_by_year_prof=[professions_year_vacancies[i] for i in professions_year_vacancies],
                        city_salary=city_salary,
                        city_vacancies=city_vacancies,
                        file_name=graph_name)
        pdf = PdfConverter(graph_name=graph_name,
                           excel_file_name=excel_file,
                           profession=profession,
                           request=request)
        pdf.generate_pdf()

    def convert_to_param_salary(self, vacancies: List[Vacancy], comparison_param: str) -> (List[ParamSalary]):
        """
        Конвертирует список вакансий по параметру сравнения в список класса ParamSalary
        @param vacancies: Набор вакансий
        :type (str)
        @param comparison_param: Параметр сравнения
        :type (str)
        @return: Список данных класса ParamSalary
        :type (List[ParamSalary])
        """
        param_salary = {}
        for vacancy in vacancies:
            dict_comparison_params = {"year": vacancy.year, "city": vacancy.area_name}
            param = dict_comparison_params[comparison_param]
            if not param_salary.__contains__(param):
                param_salary[param] = ParamSalary(param, vacancy.salary)
            else:
                param_salary[param] = param_salary[param].add_salary(vacancy.salary)
        return [param_salary[d] for d in param_salary]

    def __convert_from_param_salary_to_dict(self, param_salary: List[ParamSalary]) -> (Dict[int, int], Dict[int, int]):
        """
        Нужен для обработки списка данных класса ParamSalary и возвращения 2 словарей 1 - выбранный параметр: средняя зарплата 2 - выбранный параметр: количество вакансий
        @param param_salary: Список данных класса ParamSalary
        :type (List[ParamSalary])
        @return: Возвращает 2 словаря 1 - выбранный параметр: средняя зарплата 2 - выбранный параметр: количество вакансий
        :type (Dict[int, int], Dict[int, int])
        """
        return {x: y for x, y in zip([int(r.param) for r in param_salary], [0 if v.count_vacancies == 0 else int(v.salary / v.count_vacancies) for v in param_salary])},\
               {x: y for x, y in zip([int(r.param) for r in param_salary], [v.count_vacancies for v in param_salary])}

    def __add_missing_years(self, param_salary: List[ParamSalary], year_salary : List[ParamSalary]) -> List[ParamSalary]:
        """
        Добавляет года, пропущенные при выборке данных
        @param param_salary:  Список данных класса ParamSalary
        :type (List[ParamSalary])
        @param year_salary: Список данных класса ParamSalary, где param - year
        :type (List[ParamSalary])
        @return:
        :type (List[ParamSalary])
        """
        years = [i.param for i in year_salary]
        s_years = [el.param for el in param_salary]
        for y in years:
            if y not in s_years:
                param_salary.insert(int(y) - int(years[0]), ParamSalary(y, Salary("0", "0", "RUR")))
                param_salary[int(y) - int(years[0])].count_vacancies = 0
        return param_salary


InputConnect().process_data()


class TestMethods(unittest.TestCase):
    def test_convert_to_param_salary(self):
        vac = {"name": "Инженер", "salary_from": 35000.0, "salary_to": 45000.0, "salary_currency": "RUR",
               "area_name": "Moscow", "published_at": "2007-12-03T17:47:55+0300"}
        vac = Vacancy(vac)
        self.assertEqual(InputConnect().convert_to_param_salary([vac], "year")[0].param, '2007')

    def test_add_salary(self):
        self.assertEqual(ParamSalary("city", Salary(100, 200, "RUR")).add_salary(Salary(200, 300, "RUR")).salary, 400.0)
        self.assertEqual(
            ParamSalary("city", Salary(100, 200, "RUR")).add_salary(Salary(200, 300, "RUR")).count_vacancies, 2)

    def test_convert_to_rubles(self):
        self.assertEqual(Salary(100, 200, "RUR").convert_to_rubles(), 150.0)
        self.assertEqual(Salary(100, 200, "AZN").convert_to_rubles(), 5352.0)

    def test_init_vacancy(self):
        vac = {"name": "Инженер", "salary_from": 35000.0, "salary_to": 45000.0, "salary_currency": "RUR",
               "area_name": "Moscow", "published_at": "2012-12-03T17:47:55+0300"}
        vacancy = Vacancy(vac)
        self.assertEqual(vacancy.area_name, "Moscow")
        self.assertEqual(vacancy.year, '2012')

    def test_process_vacancies(self):
        dataSet = DataSet("32.txt").process_vacancies(
            ["name", "salary_from", 'salary_to', 'salary_currency', 'area_name', "published_at"],
            [["Системный аналитик", "75000.0", '95000.0', 'RUR', 'Москва', '2007-12-03T17:41:49+0300']])
        self.assertEqual(dataSet[0].name, 'Системный аналитик')
        self.assertEqual(dataSet[0].year, '2007')

    def test_init_param_salary(self):
        self.assertEqual(ParamSalary("year", Salary(1000, 2000, "USD")).salary, 90990.0)


if __name__ == "__main__":
    unittest.main()